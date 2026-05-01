"""
import_excel_history.py
-----------------------
Import historycznych danych portfela z Excela (Piko.xlsx) do tabeli portfolio_history w Supabase.
Importuje tylko daty, których jeszcze nie ma w bazie.

Mapowanie kolumn Excel → portfolio_history:
  - data (col 0)           → created_at
  - zysk na akcjach (col 1) → pnl_stocks
  - Zysk na etf (col 2)    → pnl_etfs
  - wartość ETF (col 3)    → value_etfs
  - Wartość Portfela (col 4) → total_value
  - Łączny zysk (col 5)    → total_pnl = pnl_stocks + pnl_etfs
  - value_stocks = total_value - value_etfs
  - realized_pnl_total z kolumny 5 (Łączny zysk) gdy > total_pnl → różnica

Wymagane:
  SUPABASE_SERVICE_KEY  – klucz service_role z Supabase
  Plik Excel pod ścieżką EXCEL_PATH

Uruchomienie:
  export SUPABASE_SERVICE_KEY="..."
  python scripts/import_excel_history.py
"""

import os
import sys
from datetime import datetime

import openpyxl
from supabase import create_client

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://swrpjlcnkrkzbwrmezwf.supabase.co")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "/mnt/c/Users/SG0706681/Downloads/Piko.xlsx")


def main():
    if not SUPABASE_SERVICE_KEY:
        print("BŁĄD: Ustaw zmienną SUPABASE_SERVICE_KEY")
        sys.exit(1)

    supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

    # 1. Pobierz istniejące daty z portfolio_history
    print("Pobieram istniejące rekordy z portfolio_history...")
    existing = supabase.table("portfolio_history") \
        .select("created_at") \
        .order("created_at", desc=False) \
        .execute()
    existing_dates = set()
    for row in (existing.data or []):
        existing_dates.add(row["created_at"][:10])  # "YYYY-MM-DD"
    print(f"  Istniejące daty w bazie: {len(existing_dates)}")
    if existing_dates:
        print(f"  Zakres: {min(existing_dates)} → {max(existing_dates)}")

    # 2. Wczytaj Excel
    print(f"\nWczytuję Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["archiwizacja"]

    rows_to_insert = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date_val = row[0]
        portfolio_val = row[4]

        if date_val is None or not isinstance(portfolio_val, (int, float)):
            continue

        # Parse date
        if isinstance(date_val, datetime):
            dt = date_val
        else:
            try:
                dt = datetime.fromisoformat(str(date_val))
            except ValueError:
                continue

        day_str = dt.strftime("%Y-%m-%d")

        # Pomiń jeśli już istnieje w bazie
        if day_str in existing_dates:
            skipped += 1
            continue

        pnl_stocks = float(row[1]) if isinstance(row[1], (int, float)) else 0.0
        pnl_etfs = float(row[2]) if isinstance(row[2], (int, float)) else 0.0
        value_etfs = float(row[3]) if isinstance(row[3], (int, float)) else 0.0
        total_value = float(portfolio_val)
        value_stocks = total_value - value_etfs
        total_pnl = pnl_stocks + pnl_etfs

        # Łączny zysk (kolumna 5) zawiera niezrealizowany + zrealizowany PnL
        laczny_zysk = float(row[5]) if isinstance(row[5], (int, float)) else None
        # realized = łączny zysk - niezrealizowany PnL (pnl_stocks + pnl_etfs)
        realized_pnl = round(laczny_zysk - total_pnl, 4) if laczny_zysk is not None else 0.0

        created_at = dt.strftime("%Y-%m-%dT%H:%M:%S+00:00")

        rows_to_insert.append({
            "created_at": created_at,
            "total_value": round(total_value, 4),
            "total_pnl": round(total_pnl, 4),
            "value_stocks": round(value_stocks, 4),
            "value_etfs": round(value_etfs, 4),
            "pnl_stocks": round(pnl_stocks, 4),
            "pnl_etfs": round(pnl_etfs, 4),
            "sp500_close": None,
            "realized_pnl_total": round(realized_pnl, 4),
        })
        existing_dates.add(day_str)  # deduplikacja w obrębie excela

    print(f"\n  Wierszy do importu: {len(rows_to_insert)}")
    print(f"  Pominiętych (już w bazie): {skipped}")

    if not rows_to_insert:
        print("\nNic do importu.")
        return

    # 3. Wstaw do Supabase partiami po 50
    BATCH_SIZE = 50
    inserted = 0
    for i in range(0, len(rows_to_insert), BATCH_SIZE):
        batch = rows_to_insert[i:i + BATCH_SIZE]
        result = supabase.table("portfolio_history").insert(batch).execute()
        inserted += len(batch)
        print(f"  Wstawiono {inserted}/{len(rows_to_insert)}...")

    print(f"\nGotowe! Zaimportowano {inserted} rekordów historycznych.")
    print(f"Zakres: {rows_to_insert[0]['created_at'][:10]} → {rows_to_insert[-1]['created_at'][:10]}")


if __name__ == "__main__":
    main()
